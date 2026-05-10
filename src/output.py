"""
Excel Output
============
Writes the daily picks to a color-coded Excel file.

Sheet 1 — Today's Picks:   Recommended bets at top, all other games below
Sheet 2 — Game Details:    Full model breakdown (pitchers, wRC+, exp runs, etc.)
Sheet 3 — Bet Tracker:     Pre-filled table for recording results + CLV
"""

import xlsxwriter
from datetime import date


# ---- Formatting helpers ----

def _fmt_odds(odds):
    """Format American odds: +150 or -110 or N/A"""
    if odds is None:
        return "N/A"
    return f"+{odds}" if odds > 0 else str(odds)


def _fmt_pct(prob):
    """Format probability as percentage: 54.3%"""
    if prob is None:
        return "N/A"
    return f"{prob * 100:.1f}%"


def load_existing_actuals(output_file):
    """
    Read any manually-entered actuals from the existing Bet Tracker so they
    survive a re-run of run.py.  Returns:
      {(matchup, market, pick_label): {"odds": val, "amount": val, "result": val,
                                        "pl": val, "closing": val, "clv": val, "notes": val}}
    """
    try:
        import openpyxl
    except ImportError:
        return {}

    if not output_file or not __import__("os").path.exists(output_file):
        return {}

    try:
        wb = openpyxl.load_workbook(output_file, data_only=True)
        ws = wb["Bet Tracker"]
    except Exception:
        return {}

    headers = [str(c.value).strip().lower() if c.value else "" for c in ws[2]]
    col = {h: i for i, h in enumerate(headers)}

    def _ci(name):
        return next((i for h, i in col.items() if name in h), None)

    c_game    = _ci("game")
    c_market  = _ci("market")
    c_pick    = _ci("pick")
    # Support both old format ("Actual Odds Taken" / "Actual Bet ($)") and
    # new format ("Odds Taken" / "Bet ($)" — merged column, always populated).
    c_aodds   = (next((i for h, i in col.items() if "actual odds" in h), None)
                 or next((i for h, i in col.items() if h == "odds taken"), None))
    c_abet    = (next((i for h, i in col.items() if "actual bet" in h), None)
                 or next((i for h, i in col.items() if h == "bet ($)"), None))
    c_result      = _ci("result")
    c_closing     = _ci("closing")
    c_clv         = _ci("clv")
    c_notes       = _ci("notes")
    c_placed      = _ci("bet placed")
    c_actual_pick = _ci("actual pick")

    c_date = _ci("date")

    if None in (c_game, c_market, c_pick):
        return {}

    today = __import__("datetime").date.today().strftime("%A, %B %d, %Y")

    saved = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        matchup    = row[c_game]   if c_game   is not None else None
        market     = row[c_market] if c_market is not None else None
        pick_label = row[c_pick]   if c_pick   is not None else None
        row_date   = str(row[c_date]).strip() if c_date is not None and row[c_date] else ""
        if not matchup or not market or not pick_label:
            continue
        # Only carry forward entries from today's date to prevent stale results
        # from a prior day's same-matchup game appearing in today's tracker.
        if row_date and row_date != today:
            continue
        key = (str(matchup).strip(), str(market).strip(), str(pick_label).strip())
        entry = {}
        for field, ci in [("odds", c_aodds), ("amount", c_abet), ("result", c_result),
                           ("closing", c_closing), ("clv", c_clv), ("notes", c_notes)]:
            if ci is not None and row[ci] not in (None, ""):
                entry[field] = row[ci]
        if c_placed is not None and row[c_placed] not in (None, ""):
            entry["bet_placed"] = str(row[c_placed]).strip().upper()
        if c_actual_pick is not None and row[c_actual_pick] not in (None, ""):
            entry["actual_pick"] = str(row[c_actual_pick]).strip()
        if entry:
            saved[key] = entry
    return saved


def write_picks_to_excel(picks, output_file="MLB_Picks.xlsx", games_date=None):
    """
    Write full daily analysis to Excel.

    picks:      list of game dicts produced by run.py
    games_date: date object for the games (defaults to today if not provided)
    """
    wb = xlsxwriter.Workbook(output_file)

    # ---- Define formats ----
    alert_fmt = wb.add_format({
        "bg_color": "#fff2cc", "border": 1, "align": "left",
        "font_color": "#7f6000", "bold": True, "text_wrap": True
    })
    title_fmt = wb.add_format({
        "bold": True, "font_size": 15, "font_color": "#FFFFFF",
        "bg_color": "#1a3a5c", "align": "center", "valign": "vcenter"
    })
    header_fmt = wb.add_format({
        "bold": True, "font_size": 10, "font_color": "#FFFFFF",
        "bg_color": "#2c5f8a", "align": "center", "valign": "vcenter",
        "border": 1, "text_wrap": True
    })
    section_green_fmt = wb.add_format({
        "bold": True, "bg_color": "#1e7e34", "font_color": "#FFFFFF",
        "align": "center"
    })
    section_grey_fmt = wb.add_format({
        "bold": True, "bg_color": "#5a5a5a", "font_color": "#FFFFFF",
        "align": "center"
    })
    green_cell = wb.add_format({
        "bg_color": "#c6efce", "border": 1, "align": "center",
        "bold": True, "font_color": "#276221"
    })
    red_cell = wb.add_format({
        "bg_color": "#ffc7ce", "border": 1, "align": "center",
        "font_color": "#9c0006"
    })
    normal = wb.add_format({"border": 1, "align": "center"})
    left   = wb.add_format({"border": 1, "align": "left"})
    bold_c = wb.add_format({"border": 1, "align": "center", "bold": True})
    money  = wb.add_format({"border": 1, "align": "center",
                             "num_format": "$#,##0.00"})
    pct    = wb.add_format({"border": 1, "align": "center",
                             "num_format": "0.0%"})

    today_str = (games_date or date.today()).strftime("%A, %B %d, %Y")

    # Confidence tier formats
    conf_high_fmt = wb.add_format({
        "bg_color": "#c6efce", "border": 1, "align": "center",
        "bold": True, "font_color": "#276221"
    })
    conf_med_fmt = wb.add_format({
        "bg_color": "#ffeb9c", "border": 1, "align": "center",
        "font_color": "#9c5700"
    })
    conf_low_fmt = wb.add_format({
        "bg_color": "#ffc7ce", "border": 1, "align": "center",
        "font_color": "#9c0006"
    })

    priority_fmt = wb.add_format({
        "bold": True, "font_color": "#7B4F00", "bg_color": "#FFD966",
        "border": 1, "align": "center"
    })
    fade_fmt = wb.add_format({
        "bold": True, "font_color": "#FFFFFF", "bg_color": "#E06020",
        "border": 1, "align": "center"
    })

    def _conf_fmt(tier):
        return {"High": conf_high_fmt, "Medium": conf_med_fmt}.get(tier, conf_low_fmt)

    def _is_priority(bet):
        return bool(bet.get("priority", False))

    def _is_fade(bet, game):
        return bool(bet.get("fade", False))

    status_ok_fmt = wb.add_format({
        "bg_color": "#c6efce", "border": 1, "align": "left",
        "font_color": "#276221", "bold": True
    })
    status_warn_fmt = wb.add_format({
        "bg_color": "#ffeb9c", "border": 1, "align": "left",
        "font_color": "#7f5000"
    })
    bp_taxed_fmt = wb.add_format({
        "bg_color": "#ffc7ce", "border": 1, "align": "left",
        "font_color": "#9c0006", "bold": True, "text_wrap": True
    })
    bp_normal_fmt = wb.add_format({
        "bg_color": "#ffeb9c", "border": 1, "align": "left",
        "font_color": "#7f5000", "text_wrap": True
    })
    bp_fresh_fmt = wb.add_format({
        "bg_color": "#c6efce", "border": 1, "align": "left",
        "font_color": "#276221"
    })

    def _bp_cell(game):
        """Build bullpen status string and format for Sheet 1."""
        parts = []
        for side, label in (("away", "Away"), ("home", "Home")):
            status = game.get(f"{side}_bp_status", "fresh")
            adj    = game.get(f"{side}_bp_era_adj", 0.0) or 0.0
            if status == "taxed":
                parts.append(f"{label}: TAXED (+{adj:.2f} ERA)")
            elif status == "normal":
                parts.append(f"{label}: used (+{adj:.2f} ERA)")
        if not parts:
            return "Fresh", bp_fresh_fmt
        text = "  |  ".join(parts)
        worst = max(game.get("away_bp_status", "fresh"), game.get("home_bp_status", "fresh"),
                    key=lambda s: {"fresh": 0, "normal": 1, "taxed": 2}.get(s, 0))
        fmt = bp_taxed_fmt if worst == "taxed" else bp_normal_fmt
        return text, fmt

    def _game_time_key(game):
        import re
        t = game.get("game_time_et", "")
        m = re.match(r'(\d+):(\d+)\s*(AM|PM)', t, re.IGNORECASE)
        if not m:
            return 9999
        h, mi, period = int(m.group(1)), int(m.group(2)), m.group(3).upper()
        if period == "PM" and h != 12: h += 12
        if period == "AM" and h == 12: h = 0
        return h * 60 + mi

    def _data_status(game):
        missing = []
        if not game.get("away_lineup_count"):
            missing.append("Away lineup")
        if not game.get("home_lineup_count"):
            missing.append("Home lineup")
        if game.get("umpire", "TBD") in ("TBD", "Unknown", "", None):
            missing.append("Umpire")
        if game.get("away_pitcher", "TBD") in ("TBD", None, ""):
            missing.append("Away SP")
        if game.get("home_pitcher", "TBD") in ("TBD", None, ""):
            missing.append("Home SP")
        if not missing:
            return "✓ Complete", True
        return "Missing: " + ", ".join(missing), False

    # ============================================================
    # SHEET 1 — TODAY'S PICKS
    # ============================================================
    ws1 = wb.add_worksheet("Today's Picks")
    ws1.set_zoom(90)

    # Title
    ws1.set_row(0, 28)
    ws1.merge_range("A1:Q1", f"MLB BETTING PICKS  —  {today_str}", title_fmt)

    # Headers
    headers = [
        "Time (ET)", "Matchup", "Market", "Pick",
        "Model Prob", "Model Odds", "Book Odds",
        "Edge", "Confidence", "Bet Size ($)", "Proj Score",
        "Away Starter", "Home Starter", "Platoon Alert", "Bullpen", "Data Status",
        "Priority"
    ]
    ws1.set_row(1, 30)
    for col, h in enumerate(headers):
        ws1.write(1, col, h, header_fmt)

    col_widths = [12, 24, 11, 20, 11, 11, 11, 9, 11, 12, 14, 18, 18, 32, 28, 30, 12]
    for i, w in enumerate(col_widths):
        ws1.set_column(i, i, w)

    row = 2
    sorted_picks = sorted(picks, key=_game_time_key)
    bet_games  = [g for g in sorted_picks if any(b.get("priority") for b in g.get("bets", []))]
    idle_games = [g for g in sorted_picks if not any(b.get("priority") for b in g.get("bets", []))]

    # Flatten priority bets only across all games, sort by game time
    all_bets_flat = []
    for game in bet_games:
        for bet in game["bets"]:
            if _is_priority(bet):
                all_bets_flat.append((game, bet))
    all_bets_flat.sort(key=lambda x: _game_time_key(x[0]))

    # Sort priority bets first, fade bets last within each time slot
    def _sort_key(x):
        game, bet = x
        is_pri  = _is_priority(bet)
        is_fade = _is_fade(bet, game)
        return (_game_time_key(game), 0 if is_pri else (2 if is_fade else 1))
    all_bets_flat.sort(key=_sort_key)

    # --- Recommended bets (sorted by time, priority first) ---
    ws1.set_row(row, 18)
    ws1.merge_range(row, 0, row, 16, "RECOMMENDED BETS  —  sorted by game time  (★ = Priority  |  ⚠ = Fade Watch)", section_green_fmt)
    row += 1

    if all_bets_flat:
        for game, bet in all_bets_flat:
            label    = f"{bet['team']}  {bet['bet_type_label']}".rstrip()
            conf     = game.get("confidence", "Low")
            is_pri   = _is_priority(bet)
            is_fade  = _is_fade(bet, game)
            pick_fmt = priority_fmt if is_pri else (fade_fmt if is_fade else green_cell)
            platoon_note = "  |  ".join(filter(None, [
                f"Away: {game['away_platoon_flag']}" if game.get("away_platoon_flag") else "",
                f"Home: {game['home_platoon_flag']}" if game.get("home_platoon_flag") else "",
            ]))
            ws1.write(row, 0,  game["game_time_et"],               normal)
            ws1.write(row, 1,  game["matchup"],                    left)
            ws1.write(row, 2,  bet["market"],                      normal)
            ws1.write(row, 3,  label,                              pick_fmt)
            ws1.write(row, 4,  bet["model_prob"],                  pct)
            ws1.write(row, 5,  _fmt_odds(bet["model_odds"]),       normal)
            ws1.write(row, 6,  _fmt_odds(bet["book_odds"]),        bold_c)
            ws1.write(row, 7,  bet["ev_pct"],                      pick_fmt)
            ws1.write(row, 8,  conf,                               _conf_fmt(conf))
            ws1.write(row, 9,  bet["bet_amount"],                  money)
            ws1.write(row, 10, game["proj_score"],                 normal)
            ws1.write(row, 11, game.get("away_pitcher", "TBD"),   normal)
            ws1.write(row, 12, game.get("home_pitcher", "TBD"),   normal)
            ws1.write(row, 13, platoon_note,                       alert_fmt if platoon_note else normal)
            bp_text, bp_fmt = _bp_cell(game)
            ws1.write(row, 14, bp_text, bp_fmt)
            status_text, status_ok = _data_status(game)
            ws1.write(row, 15, status_text, status_ok_fmt if status_ok else status_warn_fmt)
            if is_pri:
                ws1.write(row, 16, "★ PRIORITY", priority_fmt)
            elif is_fade:
                ws1.write(row, 16, "⚠ FADE WATCH", fade_fmt)
            else:
                ws1.write(row, 16, "—", normal)
            row += 1
    else:
        ws1.merge_range(row, 0, row, 15, "No +EV bets found today", red_cell)
        row += 1

    row += 1  # spacer

    # --- All other games ---
    ws1.set_row(row, 18)
    ws1.merge_range(row, 0, row, 14, "ALL OTHER GAMES  (no bet)", section_grey_fmt)
    row += 1

    for game in idle_games:
        conf = game.get("confidence", "Low")
        platoon_note = "  |  ".join(filter(None, [
            f"Away: {game['away_platoon_flag']}" if game.get("away_platoon_flag") else "",
            f"Home: {game['home_platoon_flag']}" if game.get("home_platoon_flag") else "",
        ]))
        status_text, status_ok = _data_status(game)
        ws1.write(row, 0,  game["game_time_et"],                        normal)
        ws1.write(row, 1,  game["matchup"],                             left)
        ws1.write(row, 2,  "—",                                         normal)
        ws1.write(row, 3,  "NO BET",                                    red_cell)
        ws1.write(row, 4,  _fmt_pct(game.get("home_win_prob")),         normal)
        ws1.write(row, 5,  _fmt_odds(game.get("home_model_odds")),      normal)
        ws1.write(row, 6,  _fmt_odds(game.get("home_ml_odds")),         normal)
        ws1.write(row, 7,  "—",                                         normal)
        ws1.write(row, 8,  conf,                                        _conf_fmt(conf))
        ws1.write(row, 9,  "—",                                         normal)
        ws1.write(row, 10, game["proj_score"],                          normal)
        ws1.write(row, 11, game.get("away_pitcher", "TBD"),            normal)
        ws1.write(row, 12, game.get("home_pitcher", "TBD"),            normal)
        ws1.write(row, 13, platoon_note,                                alert_fmt if platoon_note else normal)
        bp_text, bp_fmt = _bp_cell(game)
        ws1.write(row, 14, bp_text, bp_fmt)
        ws1.write(row, 15, status_text, status_ok_fmt if status_ok else status_warn_fmt)
        row += 1

    # ============================================================
    # SHEET 2 — GAME DETAILS
    # ============================================================
    ws2 = wb.add_worksheet("Game Details")
    ws2.set_zoom(90)

    ws2.set_row(0, 28)
    ws2.merge_range("A1:M1", f"GAME DETAILS  —  {today_str}", title_fmt)

    detail_headers = [
        "Matchup", "Venue", "Confidence",
        "Away Pitcher", "Away IP", "Away ERA est", "Away ERA act",
        "Home Pitcher", "Home IP", "Home ERA est", "Home ERA act",
        "Away Bullpen", "Away BP Status", "Away BP Adj",
        "Home Bullpen", "Home BP Status", "Home BP Adj",
        "Away wRC+", "Home wRC+",
        "Away Exp R", "Home Exp R",
        "Away Win%", "Home Win%",
        "Park Factor", "Weather", "Lineup?",
        "HP Umpire", "Ump Factor", "Book Total"
    ]
    ws2.set_row(1, 30)
    for col, h in enumerate(detail_headers):
        ws2.write(1, col, h, header_fmt)

    d_widths = [24, 22, 11, 18, 8, 13, 11, 18, 8, 13, 11, 12, 12, 10, 12, 12, 10, 11, 11, 11, 11, 11, 11, 12, 20, 10, 22, 11, 11]
    for i, w in enumerate(d_widths):
        ws2.set_column(i, i, w)

    for r, game in enumerate(picks, start=2):
        conf = game.get("confidence", "Low")
        ws2.write(r, 0,  game["matchup"],                                left)
        ws2.write(r, 1,  game.get("venue", ""),                         normal)
        ws2.write(r, 2,  conf,                                           _conf_fmt(conf))
        ws2.write(r, 3,  game.get("away_pitcher", "TBD"),                        normal)
        ws2.write(r, 4,  game.get("away_ip", 0),                               normal)
        ws2.write(r, 5,  game.get("away_era_est", "N/A"),                      normal)
        act = game.get("away_actual_era")
        ws2.write(r, 6,  f"{act:.2f}" if act is not None else "N/A",           normal)
        ws2.write(r, 7,  game.get("home_pitcher", "TBD"),                      normal)
        ws2.write(r, 8,  game.get("home_ip", 0),                               normal)
        ws2.write(r, 9,  game.get("home_era_est", "N/A"),                      normal)
        act = game.get("home_actual_era")
        ws2.write(r, 10, f"{act:.2f}" if act is not None else "N/A",           normal)
        # Away bullpen (cols 11-13)
        away_bp_status = game.get("away_bp_status", "fresh")
        away_bp_adj    = game.get("away_bp_era_adj", 0.0) or 0.0
        away_bp_sfmt   = bp_taxed_fmt if away_bp_status == "taxed" else (
                         bp_normal_fmt if away_bp_status == "normal" else bp_fresh_fmt)
        ws2.write(r, 11, game.get("away_bullpen_era", "N/A"),                  normal)
        ws2.write(r, 12, away_bp_status.capitalize(),                          away_bp_sfmt)
        ws2.write(r, 13, f"+{away_bp_adj:.2f}" if away_bp_adj else "—",       away_bp_sfmt)
        # Home bullpen (cols 14-16)
        home_bp_status = game.get("home_bp_status", "fresh")
        home_bp_adj    = game.get("home_bp_era_adj", 0.0) or 0.0
        home_bp_sfmt   = bp_taxed_fmt if home_bp_status == "taxed" else (
                         bp_normal_fmt if home_bp_status == "normal" else bp_fresh_fmt)
        ws2.write(r, 14, game.get("home_bullpen_era", "N/A"),                  normal)
        ws2.write(r, 15, home_bp_status.capitalize(),                          home_bp_sfmt)
        ws2.write(r, 16, f"+{home_bp_adj:.2f}" if home_bp_adj else "—",       home_bp_sfmt)
        ws2.write(r, 17, game.get("away_wrc_plus", 100),                       normal)
        ws2.write(r, 18, game.get("home_wrc_plus", 100),                       normal)
        ws2.write(r, 19, game.get("away_exp_runs", ""),                        normal)
        ws2.write(r, 20, game.get("home_exp_runs", ""),                        normal)
        ws2.write(r, 21, _fmt_pct(game.get("away_win_prob")),                  normal)
        ws2.write(r, 22, _fmt_pct(game.get("home_win_prob")),                  normal)
        ws2.write(r, 23, f"{game.get('park_factor', 1.0):.2f}",               normal)
        ws2.write(r, 24, game.get("weather", "N/A"),                           normal)
        away_lc = game.get("away_lineup_count", 0)
        home_lc = game.get("home_lineup_count", 0)
        lineup_str = f"Away {away_lc} / Home {home_lc}" if (away_lc or home_lc) else "None"
        ws2.write(r, 25, lineup_str, normal)
        ws2.write(r, 26, game.get("umpire", "TBD"),                            normal)
        ws2.write(r, 27, f"{game.get('umpire_factor', 1.0):.3f}",             normal)
        book_tot = game.get("book_total_line")
        ws2.write(r, 28, book_tot if book_tot is not None else "N/A",          normal)

    # ============================================================
    # SHEET 3 — BET TRACKER
    # ============================================================
    ws3 = wb.add_worksheet("Bet Tracker")
    ws3.set_zoom(90)

    ws3.set_row(0, 28)
    ws3.merge_range("A1:M1",
                    "BET TRACKER  —  Set Bet Placed Y/N. Use Actual Pick to override line (e.g. Over 8). Edit Odds/Bet to override model values. Enter W/L/P in Result.",
                    title_fmt)

    tracker_headers = [
        "Date", "Game", "Market", "Pick", "Actual Pick",
        "Bet Placed",
        "Odds Taken", "Bet ($)",
        "Result (W/L/P)", "P&L ($)",
        "Closing Line", "CLV", "Notes"
    ]
    ws3.set_row(1, 30)
    for col, h in enumerate(tracker_headers):
        ws3.write(1, col, h, header_fmt)

    t_widths = [14, 28, 11, 24, 18, 10, 12, 10, 14, 12, 13, 10, 24]
    for i, w in enumerate(t_widths):
        ws3.set_column(i, i, w)

    # Y/N dropdown for Bet Placed column (col 5), rows 3 onward
    ws3.data_validation(2, 5, 1000, 5, {
        "validate":      "list",
        "source":        ["Y", "N"],
        "input_title":   "Bet Placed?",
        "input_message": "Y = bet placed, N = skipped",
    })

    # Odds display format: shows +120 / -110 as signed integers
    odds_fmt = wb.add_format({"border": 1, "align": "center",
                               "num_format": '+#,##0;-#,##0;0'})
    pl_fmt   = wb.add_format({"border": 1, "align": "center",
                               "num_format": '+$#,##0.00;-$#,##0.00;$0.00'})

    # Preserve any actuals the user already filled in before this re-run
    existing_actuals = load_existing_actuals(output_file)

    tr = 2
    for game in bet_games:
        for bet in [b for b in game["bets"] if b.get("priority")]:
            pick_label = f"{bet['team']}  {bet['bet_type_label']}".rstrip()
            key = (str(game["matchup"]).strip(), str(bet["market"]).strip(), pick_label.strip())
            saved = existing_actuals.get(key, {})

            # Odds: actual if saved, else book odds (as integer so P&L formula works)
            eff_odds = saved.get("odds", bet["book_odds"])
            try:
                eff_odds = int(float(eff_odds))
            except (TypeError, ValueError):
                eff_odds = bet["book_odds"]

            # Bet: actual if saved, else model bet
            eff_bet = saved.get("amount", bet["bet_amount"])
            try:
                eff_bet = float(eff_bet)
            except (TypeError, ValueError):
                eff_bet = bet["bet_amount"]

            # P&L formula — col 6=G (Odds), col 7=H (Bet), col 8=I (Result)
            er = tr + 1  # Excel row (1-indexed)
            pl_formula = (
                f'=IF(I{er}="W",'
                f'IF(G{er}>0,H{er}*(G{er}/100),H{er}*(100/ABS(G{er}))),'
                f'IF(I{er}="L",-H{er},'
                f'IF(I{er}="P",0,"")))'
            )

            placed      = saved.get("bet_placed",   "Y")
            actual_pick = saved.get("actual_pick",  "")

            ws3.write(tr, 0,  today_str,                   normal)
            ws3.write(tr, 1,  game["matchup"],             left)
            ws3.write(tr, 2,  bet["market"],               normal)
            ws3.write(tr, 3,  pick_label,                  normal)
            ws3.write(tr, 4,  actual_pick,                 normal)     # Actual Pick (override)
            ws3.write(tr, 5,  placed,                      normal)     # Bet Placed
            ws3.write(tr, 6,  eff_odds,                    odds_fmt)   # Odds Taken
            ws3.write(tr, 7,  eff_bet,                     money)      # Bet ($)
            ws3.write(tr, 8,  saved.get("result",  ""),   normal)     # Result
            ws3.write_formula(tr, 9, pl_formula,           pl_fmt)     # P&L
            ws3.write(tr, 10, saved.get("closing", ""),   normal)     # Closing Line
            ws3.write(tr, 11, saved.get("clv",     ""),   normal)     # CLV
            ws3.write(tr, 12, saved.get("notes",  ""),    normal)     # Notes
            tr += 1

    wb.close()
    print(f"\n  Saved: {output_file}")
    return output_file
