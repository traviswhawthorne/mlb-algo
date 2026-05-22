"""
Push today's picks JSON to the resume-chat repo so the MLB web page can read it.
Called automatically at the end of run.py.

Also provides sync_actuals_to_web() which reads actual bet/odds entries from the
Excel Bet Tracker and patches the live picks JSON so the website shows what you
actually bet rather than the model suggestion.
"""

import os
import sys
import json
import shutil
import subprocess
import requests
import base64
from datetime import date

PICKS_DIR    = os.path.join(os.path.dirname(__file__), "data")
RESUME_REPO  = r"C:\Users\travi\OneDrive\Documents\travisCode\resume-chat"
PUBLIC_PICKS = os.path.join(RESUME_REPO, "public", "picks")
RESUME_GITHUB_REPO = "traviswhawthorne/resume-chat"


# ------------------------------------------------------------------ #
# Initial push — called at end of run.py
# ------------------------------------------------------------------ #

def push_picks():
    today = date.today().strftime("%Y-%m-%d")
    src   = os.path.join(PICKS_DIR, f"picks_{today}.json")

    if not os.path.exists(src):
        print(f"  Web push: no picks file found for {today}, skipping.")
        return

    if os.path.isdir(RESUME_REPO):
        # Local mode: copy to local repo and git push
        os.makedirs(PUBLIC_PICKS, exist_ok=True)
        dst = os.path.join(PUBLIC_PICKS, f"picks_{today}.json")
        shutil.copy2(src, dst)
        print(f"  Web push: copied picks to {dst}")
        _git_push(dst, f"picks: update {today}")
    elif os.environ.get("RESUME_REPO_TOKEN"):
        # Cloud mode: push directly to GitHub API
        repo_path = f"public/picks/picks_{today}.json"
        _github_api_push(src, repo_path, f"picks: update {today}")
    else:
        print("  Web push: RESUME_REPO not found and RESUME_REPO_TOKEN not set — skipping.")


# ------------------------------------------------------------------ #
# Sync actuals — reads Excel Bet Tracker, patches JSON, pushes
# ------------------------------------------------------------------ #

def sync_actuals_to_web(picks_excel_path=None, picks_date=None):
    """
    Read actual odds/bet amounts from the Excel Bet Tracker and write them
    into the picks JSON so the website shows what was actually wagered.

    picks_excel_path: path to MLB_Picks.xlsx (defaults to config.OUTPUT_FILE location)
    picks_date:       'YYYY-MM-DD' string (defaults to today)
    """
    today = picks_date or date.today().strftime("%Y-%m-%d")

    # Locate the Excel file — use config.OUTPUT_FILE if no path given
    if picks_excel_path is None:
        try:
            sys.path.insert(0, os.path.dirname(__file__))
            import config
            picks_excel_path = config.OUTPUT_FILE
        except Exception:
            picks_excel_path = os.path.join(os.path.dirname(__file__), "MLB_Picks.xlsx")

    actuals = _read_actuals_from_excel(picks_excel_path)
    if not actuals:
        print("  Actuals sync: no actual values found in Bet Tracker (columns F/H empty or file not found).")
        return

    # Load the local picks JSON
    local_json = os.path.join(PICKS_DIR, f"picks_{today}.json")
    if not os.path.exists(local_json):
        print(f"  Actuals sync: no picks JSON for {today}, skipping.")
        return

    with open(local_json) as f:
        picks = json.load(f)

    # Patch each bet with actual values where available
    changed = 0
    for game in picks.get("games", []):
        for bet in game.get("bets", []):
            pick_label = f"{bet['team']}  {bet['bet_type_label']}".rstrip()
            key = (str(game.get("matchup", "")).strip(),
                   str(bet["market"]).strip(),
                   pick_label.strip())
            if key in actuals:
                entry = actuals[key]
                if "odds" in entry:
                    bet["actual_odds"] = entry["odds"]
                if "amount" in entry:
                    bet["actual_amount"] = entry["amount"]
                if "bet_placed" in entry:
                    bet["bet_placed"] = entry["bet_placed"] != "N"
                if "actual_pick" in entry:
                    bet["actual_pick"] = entry["actual_pick"]
                changed += 1

    if not changed:
        print("  Actuals sync: no matching rows found between Excel and picks JSON.")
        return

    print(f"  Actuals sync: patched {changed} bet(s) with actual values.")

    # Save back to local data dir
    with open(local_json, "w") as f:
        json.dump(picks, f, indent=2)

    # Copy to resume-chat and push
    os.makedirs(PUBLIC_PICKS, exist_ok=True)
    dst = os.path.join(PUBLIC_PICKS, f"picks_{today}.json")
    shutil.copy2(local_json, dst)

    _git_push(dst, f"picks: sync actuals {today}")


# ------------------------------------------------------------------ #
# Helpers
# ------------------------------------------------------------------ #

def _read_actuals_from_excel(excel_path):
    """
    Read Bet Tracker sheet.  Returns:
      {(matchup, market, pick_label): {"odds": int, "amount": float}}

    Columns (0-indexed):
      0:Date  1:Game  2:Market  3:Pick  4:OddsTaken  5:ActualOdds  6:Bet  7:ActualBet
    """
    try:
        import openpyxl
    except ImportError:
        print("  Actuals sync: openpyxl not installed — run install_openpyxl.bat first.")
        return {}

    if not os.path.exists(excel_path):
        print(f"  Actuals sync: Excel file not found at {excel_path}")
        return {}

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb["Bet Tracker"]
    except Exception as e:
        print(f"  Actuals sync: could not open Bet Tracker sheet — {e}")
        return {}

    # Find column indices by header name (row 2, 0-indexed)
    headers = [str(c.value).strip().lower() if c.value else "" for c in ws[2]]
    col_game   = next((i for i, h in enumerate(headers) if "game" in h), 1)
    col_market = next((i for i, h in enumerate(headers) if "market" in h), 2)
    col_pick   = next((i for i, h in enumerate(headers) if "pick" in h), 3)
    # Support both old format (separate "Actual" columns) and new merged format
    col_act_odds   = (next((i for i, h in enumerate(headers) if "actual odds" in h), None)
                      or next((i for i, h in enumerate(headers) if h == "odds taken"), None))
    col_act_amount = (next((i for i, h in enumerate(headers) if "actual bet" in h), None)
                      or next((i for i, h in enumerate(headers) if h == "bet ($)"), None))
    col_placed      = next((i for i, h in enumerate(headers) if "bet placed" in h), None)
    col_actual_pick = next((i for i, h in enumerate(headers) if "actual pick" in h), None)

    if col_act_odds is None or col_act_amount is None:
        print("  Actuals: Bet Tracker columns not found — re-run run.py to regenerate the Excel.")
        return {}

    actuals = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        matchup    = row[col_game]
        market     = row[col_market]
        pick_label = row[col_pick]
        actual_odds   = row[col_act_odds]
        actual_amount = row[col_act_amount]

        if not matchup or not market or not pick_label:
            continue

        key = (str(matchup).strip(), str(market).strip(), str(pick_label).strip())
        entry = {}
        if actual_odds not in (None, ""):
            try:
                entry["odds"] = int(actual_odds)
            except (ValueError, TypeError):
                pass
        if actual_amount not in (None, ""):
            try:
                entry["amount"] = float(actual_amount)
            except (ValueError, TypeError):
                pass
        if col_placed is not None and row[col_placed] not in (None, ""):
            entry["bet_placed"] = str(row[col_placed]).strip().upper()
        if col_actual_pick is not None and row[col_actual_pick] not in (None, ""):
            entry["actual_pick"] = str(row[col_actual_pick]).strip()
        if entry:
            actuals[key] = entry

    return actuals


def _github_api_push(local_file, repo_path, commit_msg):
    """Push a file to GitHub via REST API. Works without a local repo clone."""
    token = os.environ.get("RESUME_REPO_TOKEN")
    if not token:
        print("  Web push: RESUME_REPO_TOKEN not set — skipping GitHub API push.")
        return

    headers = {
        "Authorization": f"token {token}",
        "Accept": "application/vnd.github.v3+json",
    }
    api_url = f"https://api.github.com/repos/{RESUME_GITHUB_REPO}/contents/{repo_path}"

    with open(local_file, "rb") as f:
        content_b64 = base64.b64encode(f.read()).decode()

    # Fetch current SHA so we can update an existing file
    r = requests.get(api_url, headers=headers, timeout=15)
    sha = r.json().get("sha") if r.status_code == 200 else None

    payload = {"message": commit_msg, "content": content_b64, "branch": "main"}
    if sha:
        payload["sha"] = sha

    r = requests.put(api_url, headers=headers, json=payload, timeout=30)
    if r.status_code in (200, 201):
        print(f"  Web push: pushed {os.path.basename(local_file)} via GitHub API")
    else:
        print(f"  Web push: GitHub API error {r.status_code} — {r.text[:200]}")


def _git_push(file_path, commit_msg):
    try:
        subprocess.run(["git", "-C", RESUME_REPO, "add", file_path], check=True)
        result = subprocess.run(["git", "-C", RESUME_REPO, "commit", "-m", commit_msg],
                                capture_output=True, text=True)
        if result.returncode != 0:
            if "nothing to commit" in result.stdout or "nothing to commit" in result.stderr:
                print(f"  Web push: already up to date, no push needed.")
                return
            raise subprocess.CalledProcessError(result.returncode, "git commit")
        subprocess.run(["git", "-C", RESUME_REPO, "pull", "--rebase", "-X", "theirs"], check=True)
        subprocess.run(["git", "-C", RESUME_REPO, "push"], check=True)
        print(f"  Web push: pushed to GitHub ({commit_msg})")
    except subprocess.CalledProcessError as e:
        print(f"  Web push: git error — {e}")


if __name__ == "__main__":
    # Running directly: sync actuals
    sync_actuals_to_web()
