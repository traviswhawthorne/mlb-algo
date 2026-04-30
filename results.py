#!/usr/bin/env python3
"""
MLB RESULTS CHECKER
===================
Run this the morning after to see how yesterday's picks did.

Usage:
  py results.py            <- checks most recent picks
  py results.py 2026-04-17 <- checks a specific date

Output: MLB_Results_YYYY-MM-DD.xlsx saved to Google Drive
"""

import sys
import os
import json
import requests
import xlsxwriter
from datetime import datetime
from glob import glob

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "src"))
import config

DATA_DIR = os.path.join(os.path.dirname(__file__), "data")
MLB_API  = "https://statsapi.mlb.com/api/v1"


# ------------------------------------------------------------------ #
# Load picks
# ------------------------------------------------------------------ #

def load_picks(target_date=None):
    if target_date:
        path = os.path.join(DATA_DIR, f"picks_{target_date}.json")
        if not os.path.exists(path):
            raise FileNotFoundError(f"No picks file found for {target_date}")
        with open(path) as f:
            return json.load(f)
    files = sorted(glob(os.path.join(DATA_DIR, "picks_*.json")))
    if not files:
        raise FileNotFoundError(
            "No picks files found in data/. Run run.py first.")
    with open(files[-1]) as f:
        return json.load(f)


# ------------------------------------------------------------------ #
# Fetch scores
# ------------------------------------------------------------------ #

def fetch_scores(game_pks):
    if not game_pks:
        return {}
    url    = f"{MLB_API}/schedule"
    params = {"sportId": 1, "gamePks": ",".join(str(p) for p in game_pks),
              "hydrate": "linescore,team"}
    try:
        resp = requests.get(url, params=params, timeout=15)
        resp.raise_for_status()
        data = resp.json()
    except Exception as e:
        print(f"  ERROR fetching scores: {e}")
        return {}

    scores = {}
    for date_entry in data.get("dates", []):
        for game in date_entry.get("games", []):
            pk     = game.get("gamePk")
            status = game.get("status", {}).get("abstractGameState", "")
            if status != "Final":
                continue
            ls         = game.get("linescore", {}).get("teams", {})
            away_score = ls.get("away", {}).get("runs")
            home_score = ls.get("home", {}).get("runs")
            if away_score is None or home_score is None:
                continue
            scores[pk] = {
                "away_team":  game["teams"]["away"]["team"]["name"],
                "home_team":  game["teams"]["home"]["team"]["name"],
                "away_score": int(away_score),
                "home_score": int(home_score),
            }
    return scores


# ------------------------------------------------------------------ #
# Margin + result
# ------------------------------------------------------------------ #

def get_margin_and_result(bet, away_team, away_score, home_score):
    """
    Returns (margin, result) where:
      margin  = float, positive=covered, negative=missed
      result  = 'WIN', 'LOSS', or 'PUSH'
    """
    market     = bet["market"]
    label      = bet["bet_type_label"]
    team       = bet["team"]
    total_line = bet.get("total_line")
    run_diff   = home_score - away_score   # positive = home won

    if market == "Moneyline":
        away_last = away_team.split()[-1].lower()
        team_last = team.split()[-1].lower()
        margin = (away_score - home_score) if team_last == away_last \
                 else (home_score - away_score)

    elif market == "Run Line":
        away_last     = away_team.split()[-1].lower()
        team_last     = team.split()[-1].lower()
        is_away       = team_last == away_last
        bet_score     = away_score if is_away else home_score
        opp_score     = home_score if is_away else away_score
        spread        = 1.5 if "+" in label else -1.5
        margin        = (bet_score - opp_score) + spread

    elif market == "Total":
        actual = away_score + home_score
        is_over = label.upper().startswith("O") or team.upper().startswith("O")
        margin = (actual - total_line) if is_over else (total_line - actual)
    else:
        margin = 0.0

    margin = round(margin, 1)
    result = "WIN" if margin > 0 else ("PUSH" if margin == 0 else "LOSS")
    return margin, result


def load_actual_bets(picks_excel_path):
    """
    Read the Bet Tracker sheet from the picks Excel and return actual values.
    Returns {(matchup, market, pick_label): {"odds": int, "amount": float}}
    Detects column positions by header name so old/new Excel formats both work.
    """
    try:
        import openpyxl
    except ImportError:
        return {}

    if not os.path.exists(picks_excel_path):
        return {}

    try:
        wb = openpyxl.load_workbook(picks_excel_path, data_only=True)
        ws = wb["Bet Tracker"]
    except Exception:
        return {}

    # Find column indices by header name (row 2, 0-indexed)
    headers = [str(c.value).strip().lower() if c.value else "" for c in ws[2]]
    col_game   = next((i for i, h in enumerate(headers) if "game" in h), 1)
    col_market = next((i for i, h in enumerate(headers) if "market" in h), 2)
    col_pick   = next((i for i, h in enumerate(headers) if "pick" in h), 3)
    # Support both old format (separate "Actual" columns) and new merged format
    col_act_odds   = (next((i for i, h in enumerate(headers) if "actual odds" in h), None)
                      or next((i for i, h in enumerate(headers) if h == "odds taken"), None))
    col_act_amount = (next((i for i, h in enumerate(headers) if "actual bet" in h), None)
                      or next((i for i, h in enumerate(headers) if h == "bet ($)"), None))
    col_placed = next((i for i, h in enumerate(headers) if "bet placed" in h), None)

    if col_act_odds is None or col_act_amount is None:
        print("  Actuals: Bet Tracker columns not found — re-run run.py to regenerate the Excel.")
        return {}

    actuals = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        matchup    = row[col_game]
        market     = row[col_market]
        pick_label = row[col_pick]
        actual_odds   = row[col_act_odds]
        actual_amount = row[col_act_amount]

        if not matchup or not market or not pick_label:
            continue

        key = (str(matchup).strip(), str(market).strip(), str(pick_label).strip())
        entry = {}
        if actual_odds not in (None, ""):
            try:
                entry["odds"] = int(actual_odds)
            except (ValueError, TypeError):
                pass
        if actual_amount not in (None, ""):
            try:
                entry["amount"] = float(actual_amount)
            except (ValueError, TypeError):
                pass
        if col_placed is not None and row[col_placed] not in (None, ""):
            entry["bet_placed"] = str(row[col_placed]).strip().upper()
        if entry:
            actuals[key] = entry

    return actuals


def calc_profit(result, bet_amount, book_odds):
    if result == "WIN":
        return round(bet_amount * book_odds / 100, 2) if book_odds > 0 \
               else round(bet_amount * 100 / abs(book_odds), 2)
    elif result == "LOSS":
        return -round(bet_amount, 2)
    return 0.0


# ------------------------------------------------------------------ #
# Write Excel
# ------------------------------------------------------------------ #

def write_results_excel(picks_date, rows, output_file):
    wb = xlsxwriter.Workbook(output_file)

    # ---- shared formats ----
    title_fmt = wb.add_format({
        "bold": True, "font_size": 15, "font_color": "#FFFFFF",
        "bg_color": "#1a3a5c", "align": "center", "valign": "vcenter"
    })
    header_fmt = wb.add_format({
        "bold": True, "font_size": 10, "font_color": "#FFFFFF",
        "bg_color": "#2c5f8a", "align": "center", "valign": "vcenter",
        "border": 1, "text_wrap": True
    })
    result_header_fmt = wb.add_format({
        "bold": True, "font_size": 10, "font_color": "#FFFFFF",
        "bg_color": "#1a3a5c", "align": "center", "valign": "vcenter",
        "border": 1, "text_wrap": True
    })
    green_cell  = wb.add_format({"bg_color": "#c6efce", "border": 1,
                                  "align": "center", "bold": True,
                                  "font_color": "#276221"})
    red_cell    = wb.add_format({"bg_color": "#ffc7ce", "border": 1,
                                  "align": "center", "bold": True,
                                  "font_color": "#9c0006"})
    yellow_cell = wb.add_format({"bg_color": "#ffeb9c", "border": 1,
                                  "align": "center", "font_color": "#9c5700"})
    conf_high   = wb.add_format({"bg_color": "#c6efce", "border": 1,
                                  "align": "center", "bold": True,
                                  "font_color": "#276221"})
    conf_med    = wb.add_format({"bg_color": "#ffeb9c", "border": 1,
                                  "align": "center", "font_color": "#9c5700"})
    conf_low    = wb.add_format({"bg_color": "#ffc7ce", "border": 1,
                                  "align": "center", "font_color": "#9c0006"})
    normal      = wb.add_format({"border": 1, "align": "center"})
    left        = wb.add_format({"border": 1, "align": "left"})
    bold_c      = wb.add_format({"border": 1, "align": "center", "bold": True})
    money       = wb.add_format({"border": 1, "align": "center",
                                  "num_format": "$#,##0.00"})
    money_pos   = wb.add_format({"border": 1, "align": "center",
                                  "num_format": "$#,##0.00",
                                  "font_color": "#276221", "bold": True})
    money_neg   = wb.add_format({"border": 1, "align": "center",
                                  "num_format": "$#,##0.00",
                                  "font_color": "#9c0006", "bold": True})
    pct         = wb.add_format({"border": 1, "align": "center",
                                  "num_format": "0.0%"})
    summary_fmt = wb.add_format({
        "bold": True, "bg_color": "#1a3a5c", "font_color": "#FFFFFF",
        "border": 1, "align": "center"
    })

    def _conf_fmt(tier):
        return {"High": conf_high, "Medium": conf_med}.get(tier, conf_low)

    def _result_fmt(r):
        return {"WIN": green_cell, "LOSS": red_cell}.get(r, yellow_cell)

    def _margin_fmt(m):
        if m is None:  return normal
        if m > 0:      return green_cell
        if m < 0:      return red_cell
        return yellow_cell

    def _pl_fmt(pl):
        if pl is None: return normal
        if pl > 0:     return money_pos
        if pl < 0:     return money_neg
        return money

    def _fmt_odds(o):
        if o is None: return "N/A"
        return f"+{o}" if o > 0 else str(o)

    # ============================================================
    # RESULTS SHEET
    # ============================================================
    ws = wb.add_worksheet("Results")
    ws.set_zoom(90)

    date_str = datetime.strptime(picks_date, "%Y-%m-%d").strftime(
        "%A, %B %d, %Y")

    # Title spans all columns (13 picks cols + 7 result cols = 20)
    ws.set_row(0, 28)
    ws.merge_range("A1:T1",
                   f"MLB RESULTS  —  {date_str}", title_fmt)

    # ---- Headers ----
    # Cols 0-12: same as Today's Picks
    picks_headers = [
        "Time (ET)", "Matchup", "Market", "Pick",
        "Model Prob", "Model Odds", "Book Odds",
        "Edge", "Confidence", "Bet Size ($)", "Proj Score",
        "Away Starter", "Home Starter"
    ]
    # Cols 13-19: result columns
    result_headers = [
        "Result", "Away Score", "Home Score", "Margin", "Profit / Loss"
    ]

    ws.set_row(1, 30)
    for col, h in enumerate(picks_headers):
        ws.write(1, col, h, header_fmt)
    for col, h in enumerate(result_headers, start=13):
        ws.write(1, col, h, result_header_fmt)

    # Column widths
    picks_widths  = [12, 24, 11, 20, 11, 11, 11, 9, 11, 12, 14, 18, 18]
    result_widths = [9, 13, 13, 9, 14]
    for i, w in enumerate(picks_widths + result_widths):
        ws.set_column(i, i, w)

    # ---- Sort: completed wins first, then losses, then pending ----
    done    = [r for r in rows if r["result"] != "PENDING"]
    pending = [r for r in rows if r["result"] == "PENDING"]
    done.sort(key=lambda r: r["margin"] or 0, reverse=True)

    row = 2
    for r in done + pending:
        res    = r["result"]
        margin = r["margin"]
        pl     = r["profit_loss"]
        conf   = r.get("confidence", "Low")

        ws.write(row, 0,  r["game_time_et"],                   normal)
        ws.write(row, 1,  r["matchup"],                        left)
        ws.write(row, 2,  r["market"],                         normal)
        ws.write(row, 3,  r["pick_label"],                     green_cell if res == "WIN"
                                                               else (red_cell if res == "LOSS"
                                                               else (yellow_cell if res == "PUSH"
                                                               else normal)))
        ws.write(row, 4,  r["model_prob"],                     pct)
        ws.write(row, 5,  _fmt_odds(r.get("model_odds")),      normal)
        ws.write(row, 6,  _fmt_odds(r["book_odds"]),           bold_c)
        ws.write(row, 7,  r["ev_pct"],                         normal)
        ws.write(row, 8,  conf,                                _conf_fmt(conf))
        ws.write(row, 9,  r["bet_amount"],                     money)
        ws.write(row, 10, r.get("proj_score", ""),             normal)
        ws.write(row, 11, r.get("away_pitcher", "TBD"),        normal)
        ws.write(row, 12, r.get("home_pitcher", "TBD"),        normal)
        # Result columns
        ws.write(row, 13, res,                                 _result_fmt(res))
        ws.write(row, 14, f"{r['away_team']} {r['away_score']}",  normal)
        ws.write(row, 15, f"{r['home_team']} {r['home_score']}",  normal)
        ws.write(row, 16, margin if margin is not None else "—",   _margin_fmt(margin))
        ws.write(row, 17, pl     if pl     is not None else "—",   _pl_fmt(pl))
        row += 1

    # ---- Summary row ----
    row += 1
    completed = [r for r in done if r["result"] in ("WIN", "LOSS", "PUSH")]
    wins     = sum(1 for r in completed if r["result"] == "WIN")
    losses   = sum(1 for r in completed if r["result"] == "LOSS")
    pushes   = sum(1 for r in completed if r["result"] == "PUSH")
    total_pl = sum(r["profit_loss"] for r in completed
                   if r["profit_loss"] is not None)
    staked   = sum(r["bet_amount"] for r in completed)
    roi      = (total_pl / staked * 100) if staked > 0 else 0.0

    ws.merge_range(row, 0, row, 13,
                   f"SUMMARY:  {wins}W - {losses}L - {pushes}P  |  "
                   f"Staked: ${staked:.2f}  |  ROI: {roi:+.1f}%",
                   summary_fmt)
    for col in range(14, 17):
        ws.write(row, col, "", summary_fmt)
    ws.write(row, 17, total_pl, _pl_fmt(total_pl))

    wb.close()
    print(f"  Saved: {output_file}")
    return output_file


# ------------------------------------------------------------------ #
# Main
# ------------------------------------------------------------------ #

def main():
    target_date = sys.argv[1] if len(sys.argv) > 1 else None

    print()
    print("=" * 62)
    print("  MLB RESULTS CHECKER")
    print("=" * 62)

    try:
        picks = load_picks(target_date)
    except FileNotFoundError as e:
        print(f"\n  ERROR: {e}")
        if not os.environ.get("CI"):
            input("\nPress Enter to close...")
        return

    picks_date = picks["date"]
    print(f"\n  Checking results for: {picks_date}")

    game_pks = [g["game_pk"] for g in picks["games"] if g.get("game_pk")]
    print(f"  Fetching scores for {len(game_pks)} game(s) ...")
    scores = fetch_scores(game_pks)
    print(f"  {len(scores)} final score(s) found")

    # Load actual bets from Bet Tracker if user filled them in
    actual_bets = load_actual_bets(config.OUTPUT_FILE)
    if actual_bets:
        print(f"  Actual bet overrides found: {len(actual_bets)} row(s)")

    # Build rows
    rows = []
    for game in picks["games"]:
        pk        = game.get("game_pk")
        away_team = game["away_team"]
        home_team = game["home_team"]
        matchup   = f"{away_team}  @  {home_team}"
        score     = scores.get(pk)

        for bet in game["bets"]:
            label      = bet["bet_type_label"]
            total_line = bet.get("total_line")

            pick_label = f"{bet['team']}  {bet['bet_type_label']}".rstrip()
            actual_key = (matchup.strip(), bet["market"].strip(), pick_label.strip())
            actual     = actual_bets.get(actual_key, {})
            bet_amount = actual.get("amount", bet["bet_amount"])
            book_odds  = actual.get("odds",   bet["book_odds"])

            if score:
                away_s = score["away_score"]
                home_s = score["home_score"]
                margin, result = get_margin_and_result(
                    bet, away_team, away_s, home_s)
                pl = calc_profit(result, bet_amount, book_odds)
                away_display = f"{away_team} {away_s}"
                home_display = f"{home_team} {home_s}"
            else:
                away_s = home_s = "—"
                away_display = away_team
                home_display = home_team
                margin = result = None
                pl = None
                result = "PENDING"

            # ev_pct: use saved string, or recompute from raw ev float
            ev_pct = bet.get("ev_pct") or (
                f"+{bet['ev']*100:.1f}%" if bet.get("ev") else "—"
            )

            rows.append({
                "game_time_et":  game.get("game_time_et", "—"),
                "matchup":       matchup,
                "market":        bet["market"],
                "pick_label":    f"{bet['team']}  {label}",
                "model_prob":    bet["model_prob"],
                "model_odds":    bet.get("model_odds"),
                "book_odds":     book_odds,
                "ev_pct":        ev_pct,
                "confidence":    game.get("confidence", "Low"),
                "bet_amount":    bet_amount,
                "proj_score":    game.get("proj_score", "—"),
                "away_pitcher":  game.get("away_pitcher", "TBD"),
                "home_pitcher":  game.get("home_pitcher", "TBD"),
                "result":        result,
                "away_team":     away_team,
                "away_score":    away_s,
                "home_team":     home_team,
                "home_score":    home_s,
                "margin":        margin,
                "profit_loss":   pl,
            })

    # Terminal summary
    print()
    completed = [r for r in rows if r["result"] != "PENDING"]
    wins    = sum(1 for r in completed if r["result"] == "WIN")
    losses  = sum(1 for r in completed if r["result"] == "LOSS")
    pushes  = sum(1 for r in completed if r["result"] == "PUSH")
    total_pl = sum(r["profit_loss"] for r in completed
                   if r["profit_loss"] is not None)

    print(f"  Record:  {wins}W - {losses}L - {pushes}P")
    print(f"  P&L:     ${total_pl:+.2f}")
    print()

    for r in sorted(completed, key=lambda x: x["margin"] or 0, reverse=True):
        icon = "WIN " if r["result"] == "WIN" else \
               ("PUSH" if r["result"] == "PUSH" else "LOSS")
        print(f"  [{icon}]  {r['pick_label']:<30}  "
              f"{r['away_team']} {r['away_score']} - "
              f"{r['home_team']} {r['home_score']}  "
              f"margin: {r['margin']:+.1f}  "
              f"P&L: ${r['profit_loss']:+.2f}")

    if not completed:
        print("  No completed games yet — run again after games finish.")

    # Write Excel
    print()
    out_file = config.OUTPUT_FILE.replace(
        "MLB_Picks.xlsx", f"MLB_Results_{picks_date}.xlsx")
    write_results_excel(picks_date, rows, out_file)

    print()
    print("=" * 62)
    print("  DONE — results saved to Google Drive")
    print("=" * 62)
    print()

    if not os.environ.get("CI"):
        input("Press Enter to close...")


if __name__ == "__main__":
    main()
